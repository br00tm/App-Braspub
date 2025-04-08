from venv import logger
import pandas as pd
import os
import json
import sys
from datetime import datetime, date, time
from openpyxl import load_workbook
from organizador_keywords import (obter_link_por_tipo_midia,
                                extrair_keywords_da_pagina, detectar_tipo_midia)
import argparse

def json_serial(obj):
    """
    Função para serializar objetos que não são nativamente serializáveis pelo JSON
    como datetime, date e time.
    """
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    elif isinstance(obj, time):
        return obj.isoformat()
    raise TypeError(f"Tipo não serializável: {type(obj)}")

def processar_planilha(caminho_planilha, aba_nome=None, primeira_linha=2, limite_linhas=None):
    """
    Processa a planilha Excel para extrair informações e complementá-las.
    
    Args:
        caminho_planilha: Caminho para a planilha Excel
        aba_nome: Nome da aba a ser processada (opcional)
        primeira_linha: Número da primeira linha a ser processada (começando em 1)
        limite_linhas: Número máximo de linhas a processar (opcional)
        
    Returns:
        Dict com status e resultados da operação
    """
    try:
        logger.info(f"Iniciando processamento da planilha: {caminho_planilha}")
        if not os.path.exists(caminho_planilha):
            logger.error(f"Arquivo não encontrado: {caminho_planilha}")
            return {'status': 'erro', 'mensagem': 'Arquivo não encontrado'}
        
        # Carregar planilha
        book = load_workbook(caminho_planilha, data_only=True)
        if aba_nome:
            if aba_nome not in book.sheetnames:
                logger.error(f"Aba '{aba_nome}' não encontrada na planilha")
                return {'status': 'erro', 'mensagem': f"Aba '{aba_nome}' não encontrada na planilha"}
            aba = book[aba_nome]
        else:
            aba = book.active
        
        # Definir número total de linhas a processar
        max_row = aba.max_row
        ultima_linha = min(max_row, primeira_linha + limite_linhas - 1) if limite_linhas else max_row
        
        resultados = []
        total_itens = ultima_linha - primeira_linha + 1
        
        # Verificar se existem colunas para link web imagem e texto
        colunas = [cell.value for cell in aba[1]]
        col_link_web_imagem = None
        col_link_web_texto = None
        
        for idx, col_name in enumerate(colunas, 1):
            if col_name and 'link web' in str(col_name).lower() and 'imagem' in str(col_name).lower():
                col_link_web_imagem = idx
                logger.info(f"Encontrada coluna Link web - Imagem: {col_name} (índice {idx})")
            elif col_name and (('link web' in str(col_name).lower() and 'texto' in str(col_name).lower()) or
                              ('link materia' in str(col_name).lower())):  # Adicionar suporte para "Link Materia"
                col_link_web_texto = idx
                logger.info(f"Encontrada coluna {col_name} (índice {idx}) - tratando como Link web - Texto")
        
        for idx, row_num in enumerate(range(primeira_linha, ultima_linha + 1)):
            try:
                # Status de progresso 
                progresso = int((idx / total_itens) * 100)
                logger.info(f"Processando linha {row_num} ({progresso}%)")
                
                # Ler dados da linha
                url_base = aba.cell(row=row_num, column=1).value
                
                if not url_base:
                    logger.warning(f"URL não encontrada na linha {row_num}")
                    continue
                
                # Verificar se temos links web específicos
                link_web_imagem = aba.cell(row=row_num, column=col_link_web_imagem).value if col_link_web_imagem else None
                link_web_texto = aba.cell(row=row_num, column=col_link_web_texto).value if col_link_web_texto else None
                
                if link_web_imagem:
                    logger.info(f"Link web - Imagem na linha {row_num}: {link_web_imagem}")
                if link_web_texto:
                    logger.info(f"Link web - Texto na linha {row_num}: {link_web_texto}")
                
                # Usar valores padrão para os campos que não conseguimos extrair
                titulo = "Título não disponível"
                publicacao = "Publicação não disponível"
                data = datetime.now().strftime("%Y-%m-%d")
                
                # Para Portal, verificar se existe link_web_texto
                if link_web_texto and link_web_texto.startswith(('http://', 'https://')):
                    # Tentar buscar o link de PDF específicamente para o tipo Portal
                    portal_link = obter_link_por_tipo_midia(link_web_texto, 'Portal')
                    logger.info(f"Portal link extraído de link_web_texto: {portal_link}")
                else:
                    # Se não temos link web texto, tentar URL base para Portal
                    portal_link = obter_link_por_tipo_midia(url_base, 'Portal')
                    logger.info(f"Portal link extraído de url_base: {portal_link}")
                
                # Para Impresso, usar link_web_imagem ou processar URL para imagem
                if link_web_imagem and link_web_imagem.startswith(('http://', 'https://')):
                    # Para Impresso, usar diretamente o link_web_imagem
                    imagem_link = link_web_imagem
                    logger.info(f"Impresso link direto do link_web_imagem: {imagem_link}")
                else:
                    # Processar URL para encontrar imagem
                    imagem_link = obter_link_por_tipo_midia(url_base, 'Impresso')
                    logger.info(f"Impresso link extraído de url_base: {imagem_link}")
                
                # Para TV, processar URL para vídeo (não usar link_web_imagem)
                video_link = obter_link_por_tipo_midia(url_base, 'TV')
                logger.info(f"TV link extraído de url_base: {video_link}")
                
                # Para Rádio, processar URL para áudio (não usar link_web_imagem)
                audio_link = obter_link_por_tipo_midia(url_base, 'Rádio')
                logger.info(f"Rádio link extraído de url_base: {audio_link}")
                
                # Extrair palavras-chave
                keywords = extrair_keywords_da_pagina(url_base)
                
                # Detectar tipo de mídia
                tipo_midia = detectar_tipo_midia(url_base)
                
                # Armazenar resultados com informações detalhadas sobre os links web
                resultados.append({
                    'url': url_base,
                    'titulo': titulo,
                    'publicacao': publicacao,
                    'data': data,
                    'tipo_midia': tipo_midia,
                    'keywords': keywords,
                    'pdf': portal_link,
                    'imagem': imagem_link,
                    'video': video_link,
                    'audio': audio_link,
                    'link_web_imagem': link_web_imagem,
                    'link_web_texto': link_web_texto
                })
                
                # Atualizar células na planilha
                aba.cell(row=row_num, column=2, value=titulo)
                aba.cell(row=row_num, column=3, value=publicacao)
                aba.cell(row=row_num, column=4, value=data)
                aba.cell(row=row_num, column=5, value=tipo_midia)
                aba.cell(row=row_num, column=6, value=', '.join(keywords) if keywords else '')
                aba.cell(row=row_num, column=7, value=portal_link)
                aba.cell(row=row_num, column=8, value=imagem_link)
                aba.cell(row=row_num, column=9, value=video_link)
                aba.cell(row=row_num, column=10, value=audio_link)
                
            except Exception as e:
                logger.error(f"Erro ao processar linha {row_num}: {str(e)}", exc_info=True)
                resultados.append({
                    'url': url_base if 'url_base' in locals() else f"Linha {row_num}",
                    'erro': str(e)
                })
        
        # Salvar planilha com os resultados
        output_path = f"{os.path.splitext(caminho_planilha)[0]}_processado.xlsx"
        book.save(output_path)
        logger.info(f"Planilha processada salva em: {output_path}")
        
        return {
            'status': 'sucesso',
            'resultados': resultados,
            'arquivo_saida': output_path
        }
        
    except Exception as e:
        logger.error(f"Erro ao processar planilha: {str(e)}", exc_info=True)
        return {'status': 'erro', 'mensagem': str(e)}

def exportar_planilha(dados, caminho_saida):
    """
    Exporta os dados processados para uma planilha Excel com abas separadas por tipo de mídia.
    Mantém apenas as colunas simplificadas e as ordena:
    - Nome do Cliente
    - Data de Inclusão
    - Título da Matéria
    - Link da Matéria
    - Veículo
    - Tipo de Mídia
    
    Args:
        dados: Dicionário com dados organizados por tipo de mídia
        caminho_saida: Caminho onde a planilha será salva
        
    Returns:
        Status da operação
    """
    try:
        # Definir a ordem desejada das colunas
        ordem_colunas = [
            'Nome do Cliente',
            'Data de Inclusão',
            'Título da Matéria',
            'Link da Matéria',
            'Veículo',
            'Tipo de Mídia'
        ]
        
        # Criar um objeto ExcelWriter
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            # Para cada tipo de mídia, criar uma aba
            for tipo, registros in dados.items():
                # Converter a lista de dicionários de volta para DataFrame
                df = pd.DataFrame(registros)
                
                # Organizar as colunas na ordem desejada (apenas as que existem)
                colunas_finais = [col for col in ordem_colunas if col in df.columns]
                if colunas_finais:
                    df = df[colunas_finais]
                
                # Substituir o nome da aba por algo válido
                nome_aba = str(tipo)[:31]  # Excel limita o nome da aba a 31 caracteres
                nome_aba = nome_aba.replace('/', '_').replace('\\', '_').replace('?', '').replace('*', '')
                nome_aba = nome_aba.replace('[', '').replace(']', '').replace(':', '')
                
                # Escrever na aba
                df.to_excel(writer, sheet_name=nome_aba, index=False)
                
                # Ajustar a largura das colunas automaticamente
                worksheet = writer.sheets[nome_aba]
                
                # Dicionário para armazenar a largura máxima de cada coluna
                max_width = {}
                
                # Inicializar o dicionário com os tamanhos dos cabeçalhos
                for idx, col in enumerate(df.columns):
                    max_width[idx] = len(str(col)) + 2  # +2 para dar um pouco de espaço extra
                
                # Calcular a largura máxima para cada coluna baseada nos dados
                for idx, col in enumerate(df.columns):
                    # Converter todos os valores para string e obter o comprimento
                    column_width = max(
                        df[col].astype(str).map(len).max(),  # Maior valor dos dados
                        max_width[idx]  # Largura atual (cabeçalho)
                    )
                    
                    # Limitar a uma largura máxima razoável (opcional)
                    max_width[idx] = min(column_width + 2, 100)  # +2 para espaço e limite de 100
                
                # Aplicar as larguras às colunas
                for idx, width in max_width.items():
                    col_letter = chr(65 + idx) if idx < 26 else chr(64 + idx // 26) + chr(65 + idx % 26)
                    worksheet.column_dimensions[col_letter].width = width
        
        return {'status': 'sucesso', 'mensagem': f'Planilha salva com sucesso em {caminho_saida}'}
        
    except Exception as e:
        return {'status': 'erro', 'mensagem': str(e)}

def main():
    """
    Função principal que analisa os argumentos de linha de comando e executa as funções apropriadas.
    
    Para processar um JSON e exportar para Excel:
    python organizador.py --json arquivo.json --saida resultado.xlsx
    
    Para processar uma planilha Excel:
    python organizador.py --planilha arquivo.xlsx [--aba "Nome da Aba"] [--primeira-linha 2] [--limite-linhas 100]
    
    Returns:
        String JSON com o resultado da operação
    """
    parser = argparse.ArgumentParser(description="Organizador de dados de midia")
    
    # Grupo de argumentos mutuamente exclusivos
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--json', help='Caminho para o arquivo JSON de entrada')
    group.add_argument('--planilha', help='Caminho para a planilha Excel a ser processada')
    
    # Argumentos para exportação de JSON para Excel
    parser.add_argument('--saida', help='Caminho para o arquivo Excel de saída (para uso com --json)')
    
    # Argumentos para processamento de planilha
    parser.add_argument('--aba', help='Nome da aba da planilha a ser processada (opcional)')
    parser.add_argument('--primeira-linha', type=int, default=2, 
                       help='Número da primeira linha a processar (padrão: 2)')
    parser.add_argument('--limite-linhas', type=int,
                       help='Número máximo de linhas a processar (opcional)')
    
    args = parser.parse_args()
    
    resultado = None
    
    # Validar argumentos
    if args.json and not args.saida:
        logger.error("Argumento --saida é obrigatório quando --json é usado")
        resultado = {'status': 'erro', 'mensagem': 'Argumento --saida é obrigatório quando --json é usado'}
    elif args.json:
        # Processar JSON para Excel
        try:
            with open(args.json, 'r', encoding='utf-8') as f:
                dados = json.load(f)
            resultado = exportar_planilha(dados, args.saida)
        except Exception as e:
            logger.error(f"Erro ao processar JSON: {str(e)}", exc_info=True)
            resultado = {'status': 'erro', 'mensagem': str(e)}
    elif args.planilha:
        # Processar planilha Excel
        try:
            resultado = processar_planilha(
                args.planilha, 
                aba_nome=args.aba,
                primeira_linha=args.primeira_linha,
                limite_linhas=args.limite_linhas
            )
        except Exception as e:
            logger.error(f"Erro ao processar planilha: {str(e)}", exc_info=True)
            resultado = {'status': 'erro', 'mensagem': str(e)}
    
    # Verificar se temos um resultado válido
    if resultado is None:
        resultado = {'status': 'erro', 'mensagem': 'Nenhum processamento foi realizado'}
    
    # Para uso em linha de comando, retornar como string JSON
    if __name__ == "__main__":
        return json.dumps(resultado, default=json_serial, ensure_ascii=False)
    else:
        # Para uso como módulo, retornar o dicionário diretamente
        return resultado

if __name__ == "__main__":
    # Chamar a função principal e imprimir o resultado
    resultado = main()
    print(resultado) 